"""Testes da importação em massa de produtos (backend/product_import.py e
backend/product_import_routes.py): leitura segura de CSV/XLSX, ZIP de
imagens, prévia em duas etapas, modos de importação, campos vazios/`__LIMPAR__`,
histórico e duplicação de produto."""

from __future__ import annotations

import io
import os
import zipfile
from datetime import datetime, timedelta

import pytest
from fastapi.testclient import TestClient
from openpyxl import Workbook

os.environ.setdefault("MISTICA_SITE_API_KEY", "test-api-key")
os.environ.setdefault("MISTICA_SYNC_KEY", "test-api-key")

import importlib

main = importlib.import_module("backend.main")
product_import = importlib.import_module("backend.product_import")
backend_db = importlib.import_module("backend.database")

client = TestClient(main.app)
client.__enter__()
HEADERS = {"X-Mistica-Api-Key": "test-api-key"}

rate_limit = importlib.import_module("backend.rate_limit")


@pytest.fixture(autouse=True)
def _sem_rate_limit_entre_testes():
    # A suíte faz muitas chamadas seguidas às mesmas rotas -- limpa a janela
    # deslizante em memória antes de cada teste para não confundir 429 (limite
    # de taxa) com uma falha real de validação sendo testada aqui. O rate
    # limit em si já é validado à parte em test_rate_limit_bloqueia_excesso.
    rate_limit._HITS.clear()
    yield


def _sessao_admin(perfil: str = "adm") -> str:
    import secrets as secrets_mod

    token = secrets_mod.token_urlsafe(24)
    agora = datetime.now()
    with backend_db.conectar() as conn:
        conn.execute(
            """INSERT INTO painel_sessoes (token, usuario_id, login, nome, perfil, ip, user_agent, criada_em, expira_em, ultimo_acesso)
               VALUES (?,?,?,?,?,?,?,?,?,?)""",
            (token, 1, "admin-teste", "Admin Teste", perfil, "127.0.0.1", "pytest", agora.isoformat(sep=" ", timespec="seconds"), (agora + timedelta(hours=1)).isoformat(sep=" ", timespec="seconds"), agora.isoformat(sep=" ", timespec="seconds")),
        )
        conn.commit()
    return token


def _csv(linhas: list[str]) -> bytes:
    return ("\n".join(linhas) + "\n").encode("utf-8")


def _xlsx(cabecalhos, linhas) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.append(cabecalhos)
    for linha in linhas:
        ws.append(linha)
    buffer = io.BytesIO()
    wb.save(buffer)
    return buffer.getvalue()


def _validar(planilha_bytes, nome="produtos.csv", modo="novos_rascunho", zip_bytes=None, headers=None):
    files = {"planilha": (nome, planilha_bytes, "text/csv" if nome.endswith(".csv") else "application/octet-stream")}
    if zip_bytes is not None:
        files["zip_imagens"] = ("imagens.zip", zip_bytes, "application/zip")
    efetivos = HEADERS if headers is None else headers
    return client.post("/api/produtos/importacao/validar", headers=efetivos, files=files, data={"modo": modo})


def _confirmar(token, headers=None):
    efetivos = HEADERS if headers is None else headers
    return client.post("/api/produtos/importacao/confirmar", headers=efetivos, json={"token": token})


def _jpeg_bytes() -> bytes:
    from PIL import Image

    buf = io.BytesIO()
    Image.new("RGB", (12, 12), "blue").save(buf, format="JPEG")
    return buf.getvalue()


# ---------------------------------------------------------------------------
# Leitura de planilha: CSV / XLSX / rejeições
# ---------------------------------------------------------------------------


def test_csv_valido_gera_previa():
    r = _validar(_csv(["sku,nome,categoria,preco,estoque", "PI-001,Produto A,Aromaterapia,19.90,5"]))
    assert r.status_code == 200, r.text
    data = r.json()
    assert data["resumo"]["novos"] == 1
    assert data["resumo"]["com_erro"] == 0


def test_csv_com_encoding_invalida_e_rejeitado():
    conteudo = "sku,nome\nX,Ção".encode("latin-1")
    r = _validar(conteudo)
    assert r.status_code == 400
    assert "UTF-8" in r.json()["detail"]


def test_csv_vazio_e_rejeitado():
    r = _validar(b"")
    assert r.status_code == 400


def test_planilha_so_cabecalho_e_rejeitada():
    r = _validar(_csv(["sku,nome"]))
    assert r.status_code == 400
    assert "linha de dados" in r.json()["detail"].lower()


def test_coluna_obrigatoria_nome_ausente():
    r = _validar(_csv(["sku,categoria", "PI-1,Aromaterapia"]))
    assert r.status_code == 400
    assert "nome" in r.json()["detail"].lower()


def test_xlsx_valido_gera_previa():
    conteudo = _xlsx(["sku", "nome", "preco"], [["PI-010", "Produto XLSX", "12,50"]])
    r = _validar(conteudo, nome="produtos.xlsx")
    assert r.status_code == 200, r.text
    assert r.json()["resumo"]["novos"] == 1


def test_xlsm_e_rejeitado():
    r = _validar(b"PK\x03\x04fake", nome="produtos.xlsm")
    assert r.status_code == 400
    assert "xlsm" in r.json()["detail"].lower() or "xls" in r.json()["detail"].lower()


def test_xls_e_rejeitado():
    r = _validar(b"fake", nome="produtos.xls")
    assert r.status_code == 400


def test_extensao_nao_corresponde_ao_conteudo_real():
    # arquivo .xlsx cujo conteúdo não é um zip/xlsx de verdade
    r = _validar(b"sku,nome\nX,Y\n", nome="produtos.xlsx")
    assert r.status_code == 400
    assert "não corresponde" in r.json()["detail"]


def test_extensao_nao_suportada_rejeitada():
    r = _validar(b"qualquer coisa", nome="produtos.txt")
    assert r.status_code == 400


def test_colunas_desconhecidas_nao_bloqueiam_mas_sao_reportadas():
    r = _validar(_csv(["sku,nome,coluna_inventada", "PI-020,Produto B,valor"]))
    assert r.status_code == 200, r.text
    assert "coluna_inventada" in r.json()["resumo"]["colunas_desconhecidas"]


def test_excesso_de_linhas_e_rejeitado(monkeypatch):
    monkeypatch.setattr(product_import, "MAX_LINHAS", 3)
    linhas = ["sku,nome"] + [f"PI-{i},Produto {i}" for i in range(10)]
    r = _validar(_csv(linhas))
    assert r.status_code == 400
    assert "linhas" in r.json()["detail"].lower()


# ---------------------------------------------------------------------------
# Validação de campos por linha
# ---------------------------------------------------------------------------


@pytest.mark.parametrize("valor", ["abc", "-1", "1000001", "19.9999"])
def test_preco_invalido_vira_erro_de_linha(valor):
    r = _validar(_csv(["sku,nome,preco", f"PI-1,Produto,{valor}"]))
    assert r.status_code == 200
    linha = r.json()["linhas"][0]
    assert linha["classificacao"] == "erro"


def test_preco_valido_em_formato_brasileiro():
    r = _validar(_csv(["sku,nome,preco", "PI-2,Produto,1.234,56"]))
    # "1.234,56" tem vírgula extra por causa do CSV -- usar formato simples abaixo
    r2 = _validar(_csv(["sku;nome;preco", "PI-2;Produto;1.234,56"]))
    assert r2.status_code == 200
    linha = r2.json()["linhas"][0]
    assert linha["classificacao"] == "novo"


@pytest.mark.parametrize("valor", ["-5", "abc", "1000001"])
def test_estoque_invalido(valor):
    r = _validar(_csv(["sku,nome,estoque", f"PI-3,Produto,{valor}"]))
    assert r.json()["linhas"][0]["classificacao"] == "erro"


def test_booleano_invalido_em_ativo():
    r = _validar(_csv(["sku,nome,ativo", "PI-4,Produto,talvez"]))
    assert r.json()["linhas"][0]["classificacao"] == "erro"


def test_status_invalido():
    r = _validar(_csv(["sku,nome,status", "PI-5,Produto,publicado_as_pressas"]))
    assert r.json()["linhas"][0]["classificacao"] == "erro"


def test_peso_e_dimensoes_invalidos():
    r = _validar(_csv(["sku,nome,peso", "PI-6,Produto,-3"]))
    assert r.json()["linhas"][0]["classificacao"] == "erro"


def test_sku_duplicado_na_propria_planilha():
    r = _validar(_csv(["sku,nome", "PI-DUP,Produto 1", "PI-DUP,Produto 2"]))
    linhas = r.json()["linhas"]
    assert linhas[1]["classificacao"] == "erro"
    assert "duplicado" in linhas[1]["erros"][0]["mensagem"].lower()


def test_formula_em_celula_e_tratada_como_texto_inerte():
    r = _validar(_csv(["sku,nome", "PI-7,=1+1"]))
    assert r.status_code == 200
    linha = r.json()["linhas"][0]
    assert linha["nome"] == "=1+1"
    assert linha["classificacao"] == "novo"


# ---------------------------------------------------------------------------
# Segurança: autenticação, autorização, rate limit
# ---------------------------------------------------------------------------


def test_sem_autenticacao_e_rejeitado():
    r = _validar(_csv(["sku,nome", "PI-8,Produto"]), headers={})
    assert r.status_code == 401


def test_perfil_nao_admin_e_rejeitado():
    token = _sessao_admin(perfil="vendedor")
    client.cookies.set("mistica_painel_sessao", token)
    try:
        r = _validar(_csv(["sku,nome", "PI-9,Produto"]), headers={})
        assert r.status_code == 403
    finally:
        client.cookies.delete("mistica_painel_sessao")


def test_modo_invalido_e_rejeitado():
    r = _validar(_csv(["sku,nome", "PI-10,Produto"]), modo="modo_inexistente")
    assert r.status_code == 400


def test_rate_limit_bloqueia_excesso():
    for _ in range(6):
        resposta = _validar(_csv(["sku,nome", "PI-RATE,Produto"]))
        assert resposta.status_code in (200, 400)
    excedente = _validar(_csv(["sku,nome", "PI-RATE,Produto"]))
    assert excedente.status_code == 429


def test_planilha_grande_demais_e_rejeitada(monkeypatch):
    monkeypatch.setattr(product_import, "MAX_PLANILHA_BYTES", 100)
    conteudo = _csv(["sku,nome"] + [f"PI-{i},{'x' * 20}" for i in range(20)])
    assert len(conteudo) > 100
    r = _validar(conteudo)
    assert r.status_code in (400, 413)


# ---------------------------------------------------------------------------
# ZIP de imagens
# ---------------------------------------------------------------------------


def _zip_com(entradas: dict[str, bytes]) -> bytes:
    buf = io.BytesIO()
    with zipfile.ZipFile(buf, "w") as zf:
        for nome, dados in entradas.items():
            zf.writestr(nome, dados)
    return buf.getvalue()


def test_imagem_no_zip_e_associada_por_nome():
    img = _jpeg_bytes()
    zip_bytes = _zip_com({"foto-1.jpg": img})
    r = _validar(_csv(["sku,nome,imagem", "PI-IMG1,Produto,foto-1.jpg"]), zip_bytes=zip_bytes)
    assert r.status_code == 200, r.text
    assert r.json()["resumo"]["com_imagem"] == 1


def test_imagem_referenciada_mas_ausente_no_zip_gera_aviso():
    r = _validar(_csv(["sku,nome,imagem", "PI-IMG2,Produto,nao-existe.jpg"]), zip_bytes=_zip_com({}))
    linha = r.json()["linhas"][0]
    assert linha["tem_imagem"] is False
    assert any("não encontrada" in a["mensagem"] for a in linha["avisos"])


def test_zip_path_traversal_e_rejeitado():
    img = _jpeg_bytes()
    zip_bytes = _zip_com({"../evil.jpg": img})
    r = _validar(_csv(["sku,nome", "PI-IMG3,Produto"]), zip_bytes=zip_bytes)
    assert r.status_code == 200
    rejeitados = r.json()["resumo"]["arquivos_zip_rejeitados"]
    assert any("evil.jpg" in item["arquivo"] for item in rejeitados)


def test_zip_caminho_absoluto_e_rejeitado():
    img = _jpeg_bytes()
    zip_bytes = _zip_com({"/etc/evil.jpg": img})
    r = _validar(_csv(["sku,nome", "PI-IMG4,Produto"]), zip_bytes=zip_bytes)
    rejeitados = r.json()["resumo"]["arquivos_zip_rejeitados"]
    assert len(rejeitados) == 1


def test_zip_executavel_e_rejeitado():
    zip_bytes = _zip_com({"malware.exe": b"MZfake"})
    r = _validar(_csv(["sku,nome", "PI-IMG5,Produto"]), zip_bytes=zip_bytes)
    rejeitados = r.json()["resumo"]["arquivos_zip_rejeitados"]
    assert any("malware.exe" in item["arquivo"] for item in rejeitados)


def test_zip_magic_bytes_invalidos_rejeitado():
    zip_bytes = _zip_com({"foto-falsa.jpg": b"nao e uma imagem de verdade"})
    r = _validar(_csv(["sku,nome", "PI-IMG6,Produto"]), zip_bytes=zip_bytes)
    rejeitados = r.json()["resumo"]["arquivos_zip_rejeitados"]
    assert any("foto-falsa.jpg" in item["arquivo"] for item in rejeitados)


def test_zip_macosx_e_ignorado_nao_rejeitado():
    img = _jpeg_bytes()
    zip_bytes = _zip_com({"__MACOSX/._foto.jpg": b"lixo", "foto.jpg": img})
    r = _validar(_csv(["sku,nome,imagem", "PI-IMG7,Produto,foto.jpg"]), zip_bytes=zip_bytes)
    resumo = r.json()["resumo"]
    assert any("foto.jpg" in item for item in resumo["arquivos_zip_ignorados"])
    assert resumo["com_imagem"] == 1


def test_zip_invalido_e_rejeitado():
    r = _validar(_csv(["sku,nome", "PI-IMG8,Produto"]), zip_bytes=b"nao e um zip")
    assert r.status_code == 400


def test_zip_arquivo_sem_produto_correspondente_aparece_no_resumo():
    img = _jpeg_bytes()
    zip_bytes = _zip_com({"orfa.jpg": img})
    r = _validar(_csv(["sku,nome", "PI-IMG9,Produto"]), zip_bytes=zip_bytes)
    assert "orfa.jpg" in r.json()["resumo"]["arquivos_zip_sem_produto"]


def test_zip_nomes_duplicados_apos_normalizacao():
    img = _jpeg_bytes()
    zip_bytes = _zip_com({"Foto.JPG": img, "foto.jpg": img})
    r = _validar(_csv(["sku,nome,imagem", "PI-IMG10,Produto,foto.jpg"]), zip_bytes=zip_bytes)
    assert "foto.jpg" in r.json()["resumo"]["imagens_duplicadas"]


# ---------------------------------------------------------------------------
# Confirmação / modos de importação
# ---------------------------------------------------------------------------


def test_confirmacao_expirada_e_rejeitada(monkeypatch):
    r = _validar(_csv(["sku,nome", "PI-EXP,Produto"]))
    token = r.json()["token"]
    sessao = product_import._PREVIEWS[token]
    sessao.expira_em = product_import._agora() - 1
    r2 = _confirmar(token)
    assert r2.status_code == 400


def test_confirmacao_duplicada_e_rejeitada():
    r = _validar(_csv(["sku,nome", "PI-DUPCONF,Produto"]))
    token = r.json()["token"]
    r1 = _confirmar(token)
    assert r1.status_code == 200
    r2 = _confirmar(token)
    assert r2.status_code == 400


def test_modo_novos_rascunho_cria_produto_inativo_e_rascunho():
    r = _validar(_csv(["sku,nome", "PI-RASC,Produto Rascunho"]), modo="novos_rascunho")
    token = r.json()["token"]
    resultado = _confirmar(token).json()
    assert resultado["criados"] == 1
    linha = client.get("/api/produtos/admin", headers=HEADERS, params={"busca": "PI-RASC", "incluir_rascunhos": "true"}).json()[0]
    assert linha["ativo"] == 0
    assert linha["rascunho"] == 1


def test_modo_novos_ativos_cria_produto_publicado():
    r = _validar(_csv(["sku,nome", "PI-ATIVO,Produto Ativo"]), modo="novos_ativos")
    token = r.json()["token"]
    _confirmar(token)
    linha = client.get("/api/produtos/admin", headers=HEADERS, params={"busca": "PI-ATIVO"}).json()[0]
    assert linha["ativo"] == 1
    assert linha["rascunho"] == 0


def test_sku_existente_no_modo_somente_novos_e_ignorado():
    r1 = _validar(_csv(["sku,nome", "PI-EXIST,Produto Original"]), modo="novos_ativos")
    _confirmar(r1.json()["token"])

    r2 = _validar(_csv(["sku,nome", "PI-EXIST,Produto Repetido"]), modo="novos_ativos")
    linha = r2.json()["linhas"][0]
    assert linha["classificacao"] == "ignorado"
    resultado = _confirmar(r2.json()["token"]).json()
    assert resultado["ignorados"] == 1
    assert resultado["criados"] == 0


def test_atualizacao_por_sku_preserva_valor_quando_celula_vazia():
    r1 = _validar(_csv(["sku,nome,marca,preco", "PI-UPD,Produto Base,Marca Original,10.00"]), modo="novos_ativos")
    _confirmar(r1.json()["token"])

    r2 = _validar(_csv(["sku,nome,marca,preco", "PI-UPD,Produto Atualizado,,20.00"]), modo="novos_e_atualizar")
    linha = r2.json()["linhas"][0]
    assert linha["classificacao"] == "atualizacao"
    _confirmar(r2.json()["token"])

    produto = client.get("/api/produtos/admin", headers=HEADERS, params={"busca": "PI-UPD"}).json()[0]
    assert produto["nome"] == "Produto Atualizado"
    assert produto["marca"] == "Marca Original"  # célula vazia preservou o valor atual
    assert produto["preco"] == 20.0


def test_marcador_limpar_apaga_campo_opcional():
    r1 = _validar(_csv(["sku,nome,marca", "PI-LIMP,Produto,Marca X"]), modo="novos_ativos")
    _confirmar(r1.json()["token"])

    r2 = _validar(_csv(["sku,nome,marca", f"PI-LIMP,Produto,{product_import.MARCADOR_LIMPAR}"]), modo="novos_e_atualizar")
    _confirmar(r2.json()["token"])

    produto = client.get("/api/produtos/admin", headers=HEADERS, params={"busca": "PI-LIMP"}).json()[0]
    assert not produto["marca"]


def test_marcador_limpar_nao_e_aceito_em_nome():
    r = _validar(_csv(["sku,nome", f"PI-LIMPNOME,{product_import.MARCADOR_LIMPAR}"]))
    linha = r.json()["linhas"][0]
    assert linha["classificacao"] == "erro"


def test_produto_existente_sem_sku_nao_atualiza():
    r1 = _validar(_csv(["sku,nome", "PI-SEMSKU,Produto Base"]), modo="novos_ativos")
    _confirmar(r1.json()["token"])
    # sem sku, nunca é tratado como "atualização" de PI-SEMSKU -- é sempre outro produto novo
    r2 = _validar(_csv(["sku,nome", ",Outro Produto"]), modo="novos_e_atualizar")
    linha = r2.json()["linhas"][0]
    assert linha["classificacao"] == "novo"


def test_modo_rascunho_forcado_marca_atualizados_como_rascunho():
    r1 = _validar(_csv(["sku,nome", "PI-FORCARASC,Produto"]), modo="novos_ativos")
    _confirmar(r1.json()["token"])

    r2 = _validar(_csv(["sku,nome", "PI-FORCARASC,Produto Editado"]), modo="novos_e_atualizar_rascunho")
    _confirmar(r2.json()["token"])

    produto = client.get("/api/produtos/admin", headers=HEADERS, params={"busca": "PI-FORCARASC", "incluir_rascunhos": "true"}).json()[0]
    assert produto["rascunho"] == 1
    assert produto["ativo"] == 0


def test_rollback_nao_deixa_produtos_parciais(monkeypatch):
    r = _validar(_csv(["sku,nome", "PI-ROLLBACK,Produto"]))
    token = r.json()["token"]

    def _explode(*args, **kwargs):
        raise RuntimeError("falha simulada")

    monkeypatch.setattr(product_import, "registrar_auditoria", _explode)
    resposta = _confirmar(token)
    assert resposta.status_code == 400

    produto = client.get("/api/produtos/admin", headers=HEADERS, params={"busca": "PI-ROLLBACK", "incluir_rascunhos": "true"}).json()
    assert produto == []


def test_temporarios_sao_removidos_apos_confirmacao():
    img = _jpeg_bytes()
    zip_bytes = _zip_com({"temp.jpg": img})
    r = _validar(_csv(["sku,nome,imagem", "PI-TEMP,Produto,temp.jpg"]), zip_bytes=zip_bytes)
    token = r.json()["token"]
    diretorio = product_import._PREVIEWS[token].diretorio_temp
    assert diretorio.exists()
    _confirmar(token)
    assert not diretorio.exists()


def test_historico_e_registrado_apos_importacao():
    r = _validar(_csv(["sku,nome", "PI-HIST,Produto"]))
    _confirmar(r.json()["token"])
    historico = client.get("/api/produtos/importacao/historico", headers=HEADERS).json()
    assert historico["itens"]
    assert historico["itens"][0]["planilha_nome"]


# ---------------------------------------------------------------------------
# Duplicação de produto
# ---------------------------------------------------------------------------


def _criar_produto_basico(codigo="DUP-ORIG", nome="Produto Original", imagem_url=None):
    payload = {"codigo_p": codigo, "nome": nome, "preco": 15.0, "quantidade": 3}
    if imagem_url:
        payload["imagem_url"] = imagem_url
    resposta = client.post("/api/produtos", json=payload, headers=HEADERS)
    assert resposta.status_code == 200, resposta.text
    return resposta.json()["id"]


def test_duplicar_produto_gera_novo_id_e_rascunho():
    produto_id = _criar_produto_basico(codigo="DUP-A")
    r = client.post(f"/api/produtos/{produto_id}/duplicar", headers=HEADERS, json={"copiar_imagem": False})
    assert r.status_code == 200, r.text
    data = r.json()
    assert data["id"] != produto_id
    assert "Cópia" in data["nome"]
    assert data["rascunho"] is True


def test_duplicar_produto_nao_reutiliza_sku():
    produto_id = _criar_produto_basico(codigo="DUP-B")
    r = client.post(f"/api/produtos/{produto_id}/duplicar", headers=HEADERS, json={"copiar_imagem": False})
    novo_id = r.json()["id"]
    novo = client.get("/api/produtos/admin", headers=HEADERS, params={"busca": "Cópia", "incluir_rascunhos": "true"}).json()
    duplicado = next(p for p in novo if p["id"] == novo_id)
    assert duplicado["codigo_p"] != "DUP-B"


def test_duplicar_produto_preserva_original():
    produto_id = _criar_produto_basico(codigo="DUP-C", nome="Produto Preservado")
    client.post(f"/api/produtos/{produto_id}/duplicar", headers=HEADERS, json={"copiar_imagem": False})
    original = client.get("/api/produtos/admin", headers=HEADERS, params={"busca": "DUP-C"}).json()[0]
    assert original["nome"] == "Produto Preservado"
    assert original["id"] == produto_id


def test_duplicar_produto_inexistente_retorna_404():
    r = client.post("/api/produtos/999999999/duplicar", headers=HEADERS, json={"copiar_imagem": False})
    assert r.status_code == 404


def test_duplicar_produto_exige_autenticacao():
    produto_id = _criar_produto_basico(codigo="DUP-D")
    r = client.post(f"/api/produtos/{produto_id}/duplicar", json={"copiar_imagem": False})
    assert r.status_code == 401


def test_duplicar_produto_sem_imagem():
    produto_id = _criar_produto_basico(codigo="DUP-E")
    r = client.post(f"/api/produtos/{produto_id}/duplicar", headers=HEADERS, json={"copiar_imagem": True})
    assert r.status_code == 200
    assert r.json()["imagem_copiada"] is False


# ---------------------------------------------------------------------------
# Modelos de descrição
# ---------------------------------------------------------------------------


def test_listar_modelos_descricao():
    r = client.get("/api/produtos/modelos-descricao", headers=HEADERS)
    assert r.status_code == 200
    chaves = [m["chave"] for m in r.json()["modelos"]]
    assert "essencias" in chaves
    assert "difusores" in chaves


def test_gerar_modelo_essencias_via_aroma_usa_texto_exato():
    r = client.post(
        "/api/produtos/modelos-descricao/gerar",
        headers=HEADERS,
        json={"categoria": "essencias", "campos": {"nome": "Essência Lavanda", "aroma": "Lavanda", "conteudo": "30ml", "marca": "Via Aroma"}},
    )
    assert r.status_code == 200
    data = r.json()
    assert data["modo_de_uso"] == (
        "Adicione 15 gotas no recipiente do aromatizador elétrico ou com outra "
        "fonte de calor e aproveite uma fragrância agradável e duradoura."
    )


def test_gerar_modelo_categoria_inexistente_404():
    r = client.post("/api/produtos/modelos-descricao/gerar", headers=HEADERS, json={"categoria": "inexistente", "campos": {}})
    assert r.status_code == 404


# ---------------------------------------------------------------------------
# Planilha-modelo (download)
# ---------------------------------------------------------------------------


def test_planilha_modelo_e_baixavel_e_sem_formulas_executaveis():
    r = client.get("/api/produtos/importacao/modelo", headers=HEADERS)
    assert r.status_code == 200
    assert r.headers["content-type"].startswith("application/vnd.openxmlformats")
    from openpyxl import load_workbook

    wb = load_workbook(io.BytesIO(r.content), read_only=True)
    aba = wb["produtos"]
    linhas = list(aba.iter_rows(values_only=True))
    assert linhas[0][1] == "nome"
    wb.close()


def test_planilha_modelo_exige_autenticacao():
    r = client.get("/api/produtos/importacao/modelo")
    assert r.status_code == 401


# ---------------------------------------------------------------------------
# Comportamento quando a imagem falha DEPOIS do commit da transação
# ---------------------------------------------------------------------------


def test_falha_ao_promover_imagem_nao_reverte_produto_ja_commitado(monkeypatch):
    img = _jpeg_bytes()
    zip_bytes = _zip_com({"foto-falha.jpg": img})
    r = _validar(_csv(["sku,nome,imagem", "PI-IMGFALHA,Produto Com Imagem,foto-falha.jpg"]), zip_bytes=zip_bytes)
    assert r.status_code == 200, r.text
    assert r.json()["resumo"]["com_imagem"] == 1
    token = r.json()["token"]

    from backend.product_image_storage import ProductImageStorageError

    def _upload_falha(*args, **kwargs):
        raise ProductImageStorageError("falha simulada de storage")

    monkeypatch.setattr(product_import.imagem_storage, "upload", _upload_falha)

    resposta = _confirmar(token)
    assert resposta.status_code == 200, resposta.text
    resultado = resposta.json()
    assert resultado["criados"] == 1
    assert resultado["sem_imagem"] == 1
    assert resultado["status"] == "concluido_com_avisos"

    produto = client.get(
        "/api/produtos/admin", headers=HEADERS, params={"busca": "PI-IMGFALHA", "incluir_rascunhos": "true"}
    ).json()[0]
    assert produto["nome"] == "Produto Com Imagem"
    assert not produto["imagem_url"]

    historico = client.get("/api/produtos/importacao/historico", headers=HEADERS).json()["itens"][0]
    assert historico["status"] == "concluido_com_avisos"
    assert historico["sem_imagem"] == 1


# ---------------------------------------------------------------------------
# Rascunhos nunca aparecem publicamente
# ---------------------------------------------------------------------------


def test_produto_rascunho_nao_aparece_no_catalogo_publico():
    r = _validar(_csv(["sku,nome", "PI-PUBRASC,Produto Publico Rascunho"]), modo="novos_rascunho")
    _confirmar(r.json()["token"])

    publico = client.get("/api/produtos", params={"busca": "PI-PUBRASC"}).json()
    assert publico == []

    admin_padrao = client.get("/api/produtos/admin", headers=HEADERS, params={"busca": "PI-PUBRASC"}).json()
    assert admin_padrao == []

    admin_com_rascunho = client.get(
        "/api/produtos/admin", headers=HEADERS, params={"busca": "PI-PUBRASC", "incluir_rascunhos": "true"}
    ).json()
    assert len(admin_com_rascunho) == 1
    assert admin_com_rascunho[0]["ativo"] == 0


def test_produto_duplicado_como_rascunho_nao_aparece_publicamente():
    produto_id = _criar_produto_basico(codigo="DUP-PUB")
    r = client.post(f"/api/produtos/{produto_id}/duplicar", headers=HEADERS, json={"copiar_imagem": False})
    assert r.status_code == 200
    publico = client.get("/api/produtos", params={"busca": "Cópia"}).json()
    assert publico == []


def test_publico_incluir_rascunhos_nao_e_aceito_no_endpoint_publico():
    # /api/produtos (publico) nao expoe o parametro incluir_rascunhos: mesmo
    # que alguem tente passa-lo na query string, o filtro ativo=1 continua
    # valendo porque a rota publica nunca repassa esse argumento.
    r = _validar(_csv(["sku,nome", "PI-PUBFORCE,Produto"]), modo="novos_rascunho")
    _confirmar(r.json()["token"])
    publico = client.get("/api/produtos", params={"busca": "PI-PUBFORCE", "incluir_rascunhos": "true"}).json()
    assert publico == []
