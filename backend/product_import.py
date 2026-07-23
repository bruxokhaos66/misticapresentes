"""Importação em massa de produtos (planilha CSV/XLSX + ZIP opcional de imagens).

Fluxo em duas etapas, nunca grava nada na primeira:

1. ``validar_planilha`` lê e valida a planilha (e o ZIP, se enviado), gera uma
   prévia guardada em memória (``_PREVIEWS``) sob um token imprevisível
   (UUID4) vinculado à sessão do administrador que enviou o arquivo, com
   validade limitada (``PREVIEW_TTL_SEGUNDOS``).
2. ``confirmar_previa`` revalida o token (existe, não expirou, pertence à
   mesma sessão, ainda não foi confirmado) e só então persiste os produtos
   dentro de uma transação real do SQLite (``backend.database.conectar``,
   que já faz rollback automático em qualquer exceção).

Nenhuma linha da planilha é gravada no banco fora do passo 2. Arquivos
temporários (planilha e imagens extraídas do ZIP) ficam em um diretório
próprio por prévia, sempre apagados ao confirmar, cancelar, expirar ou falhar.
"""

from __future__ import annotations

import io
import re
import secrets
import shutil
import sqlite3
import tempfile
import threading
import time
import unicodedata
import uuid
import zipfile
from dataclasses import dataclass, field
from datetime import datetime
from decimal import Decimal, ROUND_HALF_UP, InvalidOperation
from pathlib import Path
from typing import Optional

from openpyxl import load_workbook
from openpyxl.utils.exceptions import InvalidFileException

from backend.audit import registrar_auditoria
from backend.database import conectar
from backend.logging_config import get_logger
from backend.product_code_integrity import buscar_codigo_duplicado_incluindo_inativos
from backend.product_image_storage import is_managed_by_storage
from backend.product_routes import (
    CATEGORIA_MAX,
    DESCRICAO_MAX,
    MARCA_MAX,
    NOME_MAX,
    VALOR_MAX,
    ESTOQUE_MAX,
    normalizar_codigo_produto,
)
from backend.upload_routes import (
    MAX_IMAGE_BYTES,
    imagem_storage,
    normalizar_imagem,
    validar_imagem_real,
)

logger = get_logger(__name__)

MARCADOR_LIMPAR = "__LIMPAR__"

MAX_LINHAS = 1000
MAX_COLUNAS = 24
MAX_PLANILHA_BYTES = 10 * 1024 * 1024
MAX_ZIP_BYTES = 60 * 1024 * 1024
MAX_ZIP_ARQUIVOS = 1000
MAX_ZIP_DESCOMPACTADO_BYTES = 300 * 1024 * 1024
MAX_TAXA_COMPRESSAO = 100  # descompactado/compactado -- acima disso, suspeita de zip bomb
PREVIEW_TTL_SEGUNDOS = 20 * 60
SUBCATEGORIA_MAX = 100
DESCRICAO_CURTA_MAX = 300
DESTAQUES_MAX = 2000
MODO_DE_USO_MAX = 1000
DIMENSAO_MAX = 100_000  # gramas/centímetros, limite defensivo de faixa numérica

EXTENSOES_IMAGEM_PERMITIDAS = {".jpg", ".jpeg", ".png", ".webp"}
EXTENSOES_PERIGOSAS = {
    ".exe", ".bat", ".cmd", ".com", ".msi", ".sh", ".bash", ".ps1", ".psm1",
    ".js", ".vbs", ".jar", ".php", ".py", ".pl", ".rb", ".dll", ".scr", ".app",
}

MODOS_VALIDOS = {"novos_rascunho", "novos_ativos", "novos_e_atualizar", "novos_e_atualizar_rascunho"}

# Cabeçalho da planilha (em português, conforme especificação do painel) -> campo interno.
CAMPOS_PLANILHA: dict[str, str] = {
    "sku": "codigo_p",
    "nome": "nome",
    "categoria": "categoria",
    "subcategoria": "subcategoria",
    "marca": "marca",
    "descricao_curta": "descricao_curta",
    "descricao": "descricao",
    "preco": "preco",
    "preco_promocional": "preco_promocional",
    "estoque": "quantidade",
    "peso": "peso",
    "largura": "largura",
    "altura": "altura",
    "comprimento": "comprimento",
    "destaques": "destaques",
    "modo_de_uso": "modo_de_uso",
    "imagem": "imagem",
    "ativo": "ativo",
    "status": "status",
}

CAMPOS_TEXTO_LIMPAVEIS = {
    "subcategoria", "marca", "descricao_curta", "descricao", "destaques", "modo_de_uso",
}


class ErroImportacao(Exception):
    """Erro que impede a importação inteira (arquivo inválido, prévia
    expirada, etc.). A mensagem é sempre segura para exibir ao usuário --
    nunca contém caminho interno nem stack trace."""


def _agora() -> float:
    return time.monotonic()


def _texto_seguro(valor, *, limite: int = 120) -> str:
    """Reduz qualquer valor de célula a um texto curto e imprimível, seguro
    para aparecer na prévia (nunca stack trace/caminho interno)."""
    texto = "" if valor is None else str(valor)
    texto = "".join(ch for ch in texto if ch.isprintable() or ch in "\n\t").strip()
    texto = re.sub(r"\s+", " ", texto)
    if len(texto) > limite:
        texto = texto[:limite] + "…"
    return texto


def _texto_limpo(valor, *, limite: int) -> Optional[str]:
    if valor is None:
        return None
    texto = re.sub(r"\s+", " ", str(valor)).strip()
    if not texto:
        return None
    if len(texto) > limite:
        raise ValueError(f"excede o limite de {limite} caracteres")
    return texto


def _normalizar_preco(valor, *, campo: str) -> Optional[float]:
    if valor is None:
        return None
    texto = str(valor).strip()
    if not texto:
        return None
    if isinstance(valor, (int, float)) and not isinstance(valor, bool):
        numero = Decimal(str(valor))
    else:
        # Moeda brasileira: "R$ 1.234,56" -> 1234.56; também aceita "1234.56".
        texto = re.sub(r"[Rr]\$\s*", "", texto)
        texto = texto.replace(" ", "")
        if re.match(r"^-?\d{1,3}(\.\d{3})*(,\d+)?$", texto):
            texto = texto.replace(".", "").replace(",", ".")
        elif "," in texto and "." not in texto:
            texto = texto.replace(",", ".")
        try:
            numero = Decimal(texto)
        except InvalidOperation:
            raise ValueError(f"{campo} não é um valor numérico válido")
    try:
        numero = Decimal(numero)
    except InvalidOperation:
        raise ValueError(f"{campo} não é um valor numérico válido")
    if numero < 0 or numero > Decimal(str(VALOR_MAX)):
        raise ValueError(f"{campo} deve estar entre 0 e {VALOR_MAX:.2f}")
    if numero.as_tuple().exponent < -2:
        raise ValueError(f"{campo} deve ter no máximo duas casas decimais")
    return float(numero.quantize(Decimal("0.01"), rounding=ROUND_HALF_UP))


def _normalizar_inteiro(valor, *, campo: str, minimo: int = 0, maximo: int = ESTOQUE_MAX) -> Optional[int]:
    if valor is None:
        return None
    texto = str(valor).strip()
    if not texto:
        return None
    texto = texto.replace(".", "").replace(",", "") if re.match(r"^-?\d{1,3}([.,]\d{3})*$", texto) else texto
    try:
        numero = int(Decimal(str(valor)) if isinstance(valor, (int, float)) and not isinstance(valor, bool) else Decimal(texto))
    except (InvalidOperation, ValueError):
        raise ValueError(f"{campo} não é um número inteiro válido")
    if numero < minimo or numero > maximo:
        raise ValueError(f"{campo} deve estar entre {minimo} e {maximo}")
    return numero


def _normalizar_decimal_positivo(valor, *, campo: str, maximo: float = DIMENSAO_MAX) -> Optional[float]:
    if valor is None:
        return None
    texto = str(valor).strip()
    if not texto:
        return None
    if isinstance(valor, (int, float)) and not isinstance(valor, bool):
        numero = Decimal(str(valor))
    else:
        texto = texto.replace(",", ".")
        try:
            numero = Decimal(texto)
        except InvalidOperation:
            raise ValueError(f"{campo} não é um valor numérico válido")
    if numero < 0 or numero > Decimal(str(maximo)):
        raise ValueError(f"{campo} deve estar entre 0 e {maximo}")
    return float(numero)


_VALORES_VERDADEIROS = {"1", "sim", "s", "true", "verdadeiro", "ativo", "yes", "y"}
_VALORES_FALSOS = {"0", "não", "nao", "n", "false", "falso", "inativo", "no"}


def _normalizar_booleano(valor, *, campo: str) -> Optional[bool]:
    if valor is None:
        return None
    texto = str(valor).strip().lower()
    if not texto:
        return None
    if texto in _VALORES_VERDADEIROS:
        return True
    if texto in _VALORES_FALSOS:
        return False
    raise ValueError(f"{campo} deve ser sim/não (ou verdadeiro/falso, ativo/inativo)")


def _normalizar_status(valor) -> Optional[str]:
    if valor is None:
        return None
    texto = str(valor).strip().lower()
    if not texto:
        return None
    texto = unicodedata.normalize("NFKD", texto).encode("ascii", "ignore").decode("ascii")
    if texto in {"ativo", "publicado"}:
        return "ativo"
    if texto in {"inativo", "desativado"}:
        return "inativo"
    if texto in {"rascunho", "draft"}:
        return "rascunho"
    raise ValueError("status deve ser 'ativo', 'inativo' ou 'rascunho'")


def _normalizar_nome_arquivo(valor: str) -> str:
    """Normalização segura de nome de arquivo para correspondência com o ZIP:
    só o nome-base (sem diretórios), minúsculo, unicode normalizado (NFC)."""
    nome = Path(str(valor or "").strip().replace("\\", "/")).name
    nome = unicodedata.normalize("NFC", nome).strip().lower()
    return nome


# --------------------------------------------------------------------------
# Leitura de planilha (CSV/XLSX)
# --------------------------------------------------------------------------


def _detectar_tipo_planilha(nome_arquivo: str, conteudo: bytes) -> str:
    nome = (nome_arquivo or "").lower()
    if nome.endswith(".xlsm") or nome.endswith(".xls"):
        raise ErroImportacao("Formato não suportado. Envie um arquivo .xlsx ou .csv (não .xls/.xlsm).")
    if nome.endswith(".xlsx"):
        if not conteudo.startswith(b"PK\x03\x04"):
            raise ErroImportacao("O conteúdo do arquivo não corresponde a um .xlsx válido.")
        return "xlsx"
    if nome.endswith(".csv"):
        if conteudo.startswith(b"PK\x03\x04"):
            raise ErroImportacao("O conteúdo do arquivo não corresponde a um .csv válido.")
        return "csv"
    raise ErroImportacao("Extensão de arquivo não suportada. Envie .xlsx ou .csv.")


def _ler_csv(conteudo: bytes) -> list[list[str]]:
    try:
        texto = conteudo.decode("utf-8-sig", errors="strict")
    except UnicodeDecodeError:
        raise ErroImportacao("O arquivo CSV precisa estar codificado em UTF-8.")
    if not texto.strip():
        raise ErroImportacao("A planilha está vazia.")
    delimitador = ";" if texto.split("\n", 1)[0].count(";") > texto.split("\n", 1)[0].count(",") else ","
    import csv as _csv

    leitor = _csv.reader(io.StringIO(texto), delimiter=delimitador)
    linhas = [linha for linha in leitor if any(str(c).strip() for c in linha)]
    if not linhas:
        raise ErroImportacao("A planilha está vazia.")
    return linhas


def _ler_xlsx(conteudo: bytes) -> list[list]:
    try:
        with zipfile.ZipFile(io.BytesIO(conteudo)) as zf:
            if any(nome.startswith("xl/vbaProject") for nome in zf.namelist()):
                raise ErroImportacao("Planilhas com macros (VBA) não são aceitas.")
    except zipfile.BadZipFile:
        raise ErroImportacao("O arquivo .xlsx está corrompido ou não é válido.")
    try:
        pasta = load_workbook(io.BytesIO(conteudo), read_only=True, data_only=True, keep_vba=False)
    except InvalidFileException:
        raise ErroImportacao("O arquivo .xlsx não pôde ser aberto.")
    except zipfile.BadZipFile:
        raise ErroImportacao("O arquivo .xlsx está corrompido ou não é válido.")
    try:
        if getattr(pasta, "vba_archive", None) is not None:
            raise ErroImportacao("Planilhas com macros (VBA) não são aceitas.")
        aba = pasta.active
        if aba is None:
            raise ErroImportacao("A planilha não contém nenhuma aba.")
        if (aba.max_row or 0) > MAX_LINHAS + 1:
            raise ErroImportacao(f"A planilha excede o limite de {MAX_LINHAS} linhas.")
        if (aba.max_column or 0) > MAX_COLUNAS:
            raise ErroImportacao(f"A planilha excede o limite de {MAX_COLUNAS} colunas.")
        linhas = [list(linha) for linha in aba.iter_rows(values_only=True)]
    finally:
        pasta.close()
    linhas = [linha for linha in linhas if any(str(c).strip() for c in linha if c is not None)]
    if not linhas:
        raise ErroImportacao("A planilha está vazia.")
    return linhas


def ler_planilha(nome_arquivo: str, conteudo: bytes) -> list[list]:
    if not conteudo:
        raise ErroImportacao("Arquivo vazio.")
    if len(conteudo) > MAX_PLANILHA_BYTES:
        raise ErroImportacao(f"Planilha muito grande. Limite: {MAX_PLANILHA_BYTES // (1024 * 1024)} MB.")
    tipo = _detectar_tipo_planilha(nome_arquivo, conteudo)
    linhas = _ler_csv(conteudo) if tipo == "csv" else _ler_xlsx(conteudo)
    if len(linhas) < 2:
        raise ErroImportacao("A planilha não contém nenhuma linha de dados (só o cabeçalho).")
    if len(linhas) - 1 > MAX_LINHAS:
        raise ErroImportacao(f"A planilha excede o limite de {MAX_LINHAS} linhas.")
    if linhas and len(linhas[0]) > MAX_COLUNAS:
        raise ErroImportacao(f"A planilha excede o limite de {MAX_COLUNAS} colunas.")
    return linhas


def mapear_cabecalhos(cabecalhos: list) -> tuple[dict[int, str], list[str]]:
    mapa: dict[int, str] = {}
    desconhecidas: list[str] = []
    for indice, bruto in enumerate(cabecalhos):
        chave = unicodedata.normalize("NFKD", str(bruto or "").strip().lower()).encode("ascii", "ignore").decode("ascii")
        chave = re.sub(r"\s+", "_", chave)
        if chave in CAMPOS_PLANILHA:
            mapa[indice] = CAMPOS_PLANILHA[chave]
        elif chave:
            desconhecidas.append(_texto_seguro(bruto, limite=40))
    if "nome" not in mapa.values():
        raise ErroImportacao("Coluna obrigatória ausente: 'nome'.")
    return mapa, desconhecidas


# --------------------------------------------------------------------------
# ZIP de imagens
# --------------------------------------------------------------------------


@dataclass
class ResultadoZip:
    imagens: dict[str, Path] = field(default_factory=dict)
    duplicadas: list[str] = field(default_factory=list)
    ignorados_sistema: list[str] = field(default_factory=list)
    rejeitados: list[dict] = field(default_factory=list)


def processar_zip(conteudo: bytes, destino_dir: Path) -> ResultadoZip:
    if len(conteudo) > MAX_ZIP_BYTES:
        raise ErroImportacao(f"ZIP muito grande. Limite: {MAX_ZIP_BYTES // (1024 * 1024)} MB.")
    resultado = ResultadoZip()
    try:
        zf = zipfile.ZipFile(io.BytesIO(conteudo))
    except zipfile.BadZipFile:
        raise ErroImportacao("O arquivo ZIP está corrompido ou não é válido.")
    with zf:
        infos = zf.infolist()
        if len(infos) > MAX_ZIP_ARQUIVOS:
            raise ErroImportacao(f"O ZIP excede o limite de {MAX_ZIP_ARQUIVOS} arquivos.")
        total_descompactado = sum(info.file_size for info in infos)
        if total_descompactado > MAX_ZIP_DESCOMPACTADO_BYTES:
            raise ErroImportacao("O ZIP excede o limite de tamanho descompactado.")

        vistos: dict[str, int] = {}
        for info in infos:
            nome_bruto = info.filename
            if info.is_dir():
                continue
            base = Path(nome_bruto.replace("\\", "/")).name
            if not base:
                continue
            if "__MACOSX" in nome_bruto or base.startswith("."):
                resultado.ignorados_sistema.append(_texto_seguro(base, limite=80))
                continue
            if nome_bruto.startswith("/") or nome_bruto.startswith("\\") or ".." in Path(nome_bruto).parts or re.match(r"^[a-zA-Z]:", nome_bruto):
                resultado.rejeitados.append({"arquivo": _texto_seguro(base, limite=80), "motivo": "caminho não permitido"})
                continue
            modo_unix = (info.external_attr >> 16) & 0xFFFF
            eh_link_simbolico = bool(modo_unix) and (modo_unix & 0o170000) == 0o120000
            if eh_link_simbolico:
                resultado.rejeitados.append({"arquivo": _texto_seguro(base, limite=80), "motivo": "link simbólico não permitido"})
                continue
            ext = Path(base).suffix.lower()
            if ext in EXTENSOES_PERIGOSAS:
                resultado.rejeitados.append({"arquivo": _texto_seguro(base, limite=80), "motivo": "tipo de arquivo não permitido"})
                continue
            if ext not in EXTENSOES_IMAGEM_PERMITIDAS:
                resultado.rejeitados.append({"arquivo": _texto_seguro(base, limite=80), "motivo": "extensão de imagem não suportada"})
                continue
            if info.file_size > MAX_IMAGE_BYTES:
                resultado.rejeitados.append({"arquivo": _texto_seguro(base, limite=80), "motivo": "imagem excede o tamanho máximo"})
                continue
            taxa = info.file_size / max(info.compress_size, 1)
            if taxa > MAX_TAXA_COMPRESSAO:
                resultado.rejeitados.append({"arquivo": _texto_seguro(base, limite=80), "motivo": "taxa de compressão suspeita"})
                continue

            nome_normalizado = _normalizar_nome_arquivo(base)
            vistos[nome_normalizado] = vistos.get(nome_normalizado, 0) + 1
            if vistos[nome_normalizado] > 1:
                if nome_normalizado not in resultado.duplicadas:
                    resultado.duplicadas.append(nome_normalizado)
                resultado.imagens.pop(nome_normalizado, None)
                continue

            try:
                dados = zf.read(info)
            except (zipfile.BadZipFile, RuntimeError, OSError):
                resultado.rejeitados.append({"arquivo": _texto_seguro(base, limite=80), "motivo": "falha ao extrair arquivo"})
                continue
            if len(dados) != info.file_size or len(dados) > MAX_IMAGE_BYTES:
                resultado.rejeitados.append({"arquivo": _texto_seguro(base, limite=80), "motivo": "tamanho inconsistente"})
                continue
            assinatura_ok = (
                dados.startswith(b"\xff\xd8\xff") if ext in (".jpg", ".jpeg")
                else dados.startswith(b"\x89PNG\r\n\x1a\n") if ext == ".png"
                else dados[:4] == b"RIFF" and dados[8:12] == b"WEBP" if ext == ".webp"
                else False
            )
            if not assinatura_ok:
                resultado.rejeitados.append({"arquivo": _texto_seguro(base, limite=80), "motivo": "conteúdo não corresponde à extensão declarada"})
                continue

            destino = destino_dir / f"{uuid.uuid4().hex}{ext}"
            destino.write_bytes(dados)
            resultado.imagens[nome_normalizado] = destino
    return resultado


# --------------------------------------------------------------------------
# Validação/classificação de linhas
# --------------------------------------------------------------------------


@dataclass
class LinhaImportacao:
    numero: int
    dados: dict = field(default_factory=dict)
    erros: list[dict] = field(default_factory=list)
    avisos: list[dict] = field(default_factory=list)
    classificacao: str = "novo"  # novo | atualizacao | ignorado | erro
    produto_id_existente: Optional[int] = None
    imagem_nome: Optional[str] = None
    tem_imagem: bool = False


CAMPOS_TEXTO_SIMPLES = {
    "categoria": CATEGORIA_MAX,
    "subcategoria": SUBCATEGORIA_MAX,
    "marca": MARCA_MAX,
    "descricao_curta": DESCRICAO_CURTA_MAX,
    "descricao": DESCRICAO_MAX,
    "destaques": DESTAQUES_MAX,
    "modo_de_uso": MODO_DE_USO_MAX,
}


def _processar_valor_campo(chave: str, valor_bruto, linha: LinhaImportacao, cabecalho_exibicao: str):
    try:
        if valor_bruto is not None and str(valor_bruto).strip() == MARCADOR_LIMPAR:
            if chave not in CAMPOS_TEXTO_LIMPAVEIS:
                linha.erros.append({
                    "coluna": cabecalho_exibicao, "valor": MARCADOR_LIMPAR,
                    "mensagem": f"'{MARCADOR_LIMPAR}' não pode ser usado no campo '{cabecalho_exibicao}'.",
                })
                return
            linha.dados[chave] = MARCADOR_LIMPAR
            return
        if chave == "codigo_p":
            linha.dados[chave] = normalizar_codigo_produto(valor_bruto)
        elif chave == "nome":
            linha.dados[chave] = _texto_limpo(valor_bruto, limite=NOME_MAX)
        elif chave in CAMPOS_TEXTO_SIMPLES:
            linha.dados[chave] = _texto_limpo(valor_bruto, limite=CAMPOS_TEXTO_SIMPLES[chave])
        elif chave in ("preco", "preco_promocional"):
            linha.dados[chave] = _normalizar_preco(valor_bruto, campo=cabecalho_exibicao)
        elif chave == "quantidade":
            linha.dados[chave] = _normalizar_inteiro(valor_bruto, campo=cabecalho_exibicao, maximo=ESTOQUE_MAX)
        elif chave in ("peso", "largura", "altura", "comprimento"):
            linha.dados[chave] = _normalizar_decimal_positivo(valor_bruto, campo=cabecalho_exibicao)
        elif chave == "ativo":
            linha.dados[chave] = _normalizar_booleano(valor_bruto, campo=cabecalho_exibicao)
        elif chave == "status":
            linha.dados[chave] = _normalizar_status(valor_bruto)
        elif chave == "imagem":
            texto = _texto_limpo(valor_bruto, limite=180)
            linha.dados[chave] = texto
        else:
            linha.dados[chave] = valor_bruto
    except ValueError as exc:
        linha.erros.append({
            "coluna": cabecalho_exibicao,
            "valor": _texto_seguro(valor_bruto, limite=60),
            "mensagem": f"{exc}. Corrija o valor e reenvie a planilha.",
        })


def construir_previa(
    conn,
    linhas_brutas: list[list],
    mapa_cabecalhos: dict[int, str],
    modo: str,
    zip_resultado: Optional[ResultadoZip],
) -> tuple[list[LinhaImportacao], dict]:
    indice_para_nome = {i: nome for i, nome in mapa_cabecalhos.items()}
    linhas: list[LinhaImportacao] = []
    skus_na_planilha: dict[str, int] = {}
    imagens_usadas: set[str] = set()

    for numero, bruta in enumerate(linhas_brutas[1:], start=2):
        linha = LinhaImportacao(numero=numero)
        for indice, chave in indice_para_nome.items():
            valor = bruta[indice] if indice < len(bruta) else None
            _processar_valor_campo(chave, valor, linha, chave)

        if linha.erros:
            linha.classificacao = "erro"
            linhas.append(linha)
            continue

        nome = linha.dados.get("nome")
        codigo = linha.dados.get("codigo_p")

        if codigo:
            if codigo in skus_na_planilha:
                linha.erros.append({
                    "coluna": "sku", "valor": codigo,
                    "mensagem": f"SKU duplicado dentro da própria planilha (já usado na linha {skus_na_planilha[codigo]}).",
                })
                linha.classificacao = "erro"
                linhas.append(linha)
                continue
            skus_na_planilha[codigo] = numero

        existente = buscar_codigo_duplicado_incluindo_inativos(conn, codigo) if codigo else None
        linha.produto_id_existente = int(existente["id"]) if existente else None

        if not existente and not nome:
            linha.erros.append({"coluna": "nome", "valor": "", "mensagem": "Nome é obrigatório para criar um novo produto."})
            linha.classificacao = "erro"
            linhas.append(linha)
            continue

        pode_atualizar = modo in ("novos_e_atualizar", "novos_e_atualizar_rascunho")
        if existente:
            if not codigo:
                linha.erros.append({"coluna": "sku", "valor": "", "mensagem": "Produto existente exige SKU para ser atualizado."})
                linha.classificacao = "erro"
            elif not pode_atualizar:
                linha.classificacao = "ignorado"
                linha.avisos.append({"coluna": "sku", "valor": codigo, "mensagem": "Já existe um produto com este SKU; este modo não atualiza produtos existentes."})
            else:
                linha.classificacao = "atualizacao"
        else:
            linha.classificacao = "novo"

        nome_imagem = linha.dados.get("imagem")
        if nome_imagem:
            normalizado = _normalizar_nome_arquivo(nome_imagem)
            linha.imagem_nome = normalizado
            if zip_resultado and normalizado in zip_resultado.imagens:
                linha.tem_imagem = True
                imagens_usadas.add(normalizado)
            elif zip_resultado and normalizado in zip_resultado.duplicadas:
                linha.avisos.append({"coluna": "imagem", "valor": nome_imagem, "mensagem": "Nome de imagem duplicado no ZIP; nenhuma das cópias foi usada."})
            else:
                linha.avisos.append({"coluna": "imagem", "valor": nome_imagem, "mensagem": "Imagem referenciada não encontrada no ZIP enviado."})

        linhas.append(linha)

    arquivos_zip_sem_produto = []
    if zip_resultado:
        arquivos_zip_sem_produto = sorted(set(zip_resultado.imagens.keys()) - imagens_usadas)

    resumo = {
        "total_linhas": len(linhas),
        "validas": sum(1 for l in linhas if l.classificacao != "erro"),
        "com_erro": sum(1 for l in linhas if l.classificacao == "erro"),
        "novos": sum(1 for l in linhas if l.classificacao == "novo"),
        "atualizacoes": sum(1 for l in linhas if l.classificacao == "atualizacao"),
        "ignorados": sum(1 for l in linhas if l.classificacao == "ignorado"),
        "sem_imagem": sum(1 for l in linhas if l.classificacao in ("novo", "atualizacao") and not l.tem_imagem and not (l.produto_id_existente and not l.dados.get("imagem"))),
        "com_imagem": sum(1 for l in linhas if l.tem_imagem),
        "imagens_duplicadas": zip_resultado.duplicadas if zip_resultado else [],
        "arquivos_zip_sem_produto": arquivos_zip_sem_produto,
        "arquivos_zip_ignorados": zip_resultado.ignorados_sistema if zip_resultado else [],
        "arquivos_zip_rejeitados": zip_resultado.rejeitados if zip_resultado else [],
    }
    return linhas, resumo


# --------------------------------------------------------------------------
# Sessão de prévia (em memória, com TTL e vínculo à sessão do admin)
# --------------------------------------------------------------------------


@dataclass
class PreviewSession:
    token: str
    dono: str
    admin_login: str
    modo: str
    planilha_nome: str
    zip_nome: Optional[str]
    linhas: list[LinhaImportacao]
    resumo: dict
    diretorio_temp: Path
    zip_resultado: Optional[ResultadoZip]
    criado_em: float
    expira_em: float
    confirmado: bool = False


_PREVIEWS: dict[str, PreviewSession] = {}
_LOCK = threading.Lock()


def _limpar_diretorio(caminho: Path) -> None:
    try:
        shutil.rmtree(caminho, ignore_errors=True)
    except Exception:
        pass


def _expirar_previas() -> None:
    agora = _agora()
    expirados = [token for token, sessao in _PREVIEWS.items() if sessao.expira_em < agora]
    for token in expirados:
        sessao = _PREVIEWS.pop(token, None)
        if sessao:
            _limpar_diretorio(sessao.diretorio_temp)


def criar_previa(*, dono: str, admin_login: str, modo: str, planilha_nome: str, zip_nome: Optional[str],
                  linhas: list[LinhaImportacao], resumo: dict, diretorio_temp: Path,
                  zip_resultado: Optional[ResultadoZip]) -> PreviewSession:
    token = uuid.uuid4().hex + secrets.token_hex(8)
    agora = _agora()
    sessao = PreviewSession(
        token=token, dono=dono, admin_login=admin_login, modo=modo,
        planilha_nome=planilha_nome, zip_nome=zip_nome, linhas=linhas, resumo=resumo,
        diretorio_temp=diretorio_temp, zip_resultado=zip_resultado,
        criado_em=agora, expira_em=agora + PREVIEW_TTL_SEGUNDOS,
    )
    with _LOCK:
        _expirar_previas()
        _PREVIEWS[token] = sessao
    return sessao


def obter_previa(token: str, *, dono: str) -> PreviewSession:
    with _LOCK:
        _expirar_previas()
        sessao = _PREVIEWS.get(token)
        if not sessao:
            raise ErroImportacao("Prévia não encontrada ou expirada. Envie a planilha novamente.")
        if sessao.dono != dono:
            raise ErroImportacao("Esta prévia pertence a outra sessão administrativa.")
        return sessao


def descartar_previa(token: str) -> None:
    with _LOCK:
        sessao = _PREVIEWS.pop(token, None)
    if sessao:
        _limpar_diretorio(sessao.diretorio_temp)


def novo_diretorio_temp() -> Path:
    return Path(tempfile.mkdtemp(prefix="mistica_import_"))


def limpar_diretorio_temp(caminho: Path) -> None:
    _limpar_diretorio(caminho)


# --------------------------------------------------------------------------
# Confirmação / persistência
# --------------------------------------------------------------------------


CAMPOS_PRODUTO_ATUALIZAVEIS = [
    "nome", "marca", "preco", "quantidade", "categoria", "subcategoria", "descricao_curta",
    "descricao", "preco_promocional", "peso", "largura", "altura", "comprimento", "destaques", "modo_de_uso",
]


def _valor_final_para_update(existente: dict, chave: str, novo_valor):
    """Célula vazia (``None``) preserva o valor atual; ``__LIMPAR__`` apaga
    campos opcionais permitidos; qualquer outro valor substitui."""
    if novo_valor is None:
        return existente.get(chave)
    if novo_valor == MARCADOR_LIMPAR:
        return None
    return novo_valor


def _estado_ativo_rascunho(modo: str, linha: LinhaImportacao, existente: Optional[dict]) -> tuple[int, int]:
    status = linha.dados.get("status")
    ativo_col = linha.dados.get("ativo")

    if modo == "novos_rascunho":
        if existente:
            return int(existente.get("ativo") or 0), int(existente.get("rascunho") or 0)
        return 0, 1
    if modo == "novos_ativos":
        if existente:
            return int(existente.get("ativo") or 0), int(existente.get("rascunho") or 0)
        return 1, 0
    if modo == "novos_e_atualizar_rascunho":
        return 0, 1

    # modo == "novos_e_atualizar": respeita status/ativo quando informado;
    # célula vazia preserva o status existente (ou usa rascunho seguro para novos).
    if status == "ativo":
        return 1, 0
    if status == "inativo":
        return 0, 0
    if status == "rascunho":
        return 0, 1
    if ativo_col is True:
        return 1, 0
    if ativo_col is False:
        return 0, 0
    if existente:
        return int(existente.get("ativo") or 0), int(existente.get("rascunho") or 0)
    return 0, 1


def confirmar_previa(token: str, *, dono: str, admin_login: str) -> dict:
    sessao = obter_previa(token, dono=dono)
    with _LOCK:
        if sessao.confirmado:
            raise ErroImportacao("Esta importação já foi confirmada anteriormente.")
        sessao.confirmado = True

    inicio = datetime.now()
    import_id = str(uuid.uuid4())
    criados = atualizados = ignorados = com_erro = sem_imagem = 0
    falhou = False
    mensagem_erro = None
    pendentes_imagem: list[tuple[int, Path, str]] = []

    try:
        with conectar() as conn:
            for linha in sessao.linhas:
                if linha.classificacao == "erro":
                    com_erro += 1
                    continue
                if linha.classificacao == "ignorado":
                    ignorados += 1
                    continue

                codigo = linha.dados.get("codigo_p")
                existente_row = None
                if linha.produto_id_existente:
                    existente_row = conn.execute("SELECT * FROM produtos WHERE id=?", (linha.produto_id_existente,)).fetchone()
                    existente_row = dict(existente_row) if existente_row else None

                ativo, rascunho = _estado_ativo_rascunho(sessao.modo, linha, existente_row)
                agora_iso = datetime.now().isoformat(timespec="seconds")

                if existente_row:
                    valores = {
                        campo: _valor_final_para_update(existente_row, campo, linha.dados.get(campo))
                        for campo in CAMPOS_PRODUTO_ATUALIZAVEIS
                    }
                    custo = float(existente_row.get("custo") or 0.0)
                    preco = float(valores["preco"] if valores["preco"] is not None else existente_row.get("preco") or 0.0)
                    lucro = float(Decimal(str(preco)) - Decimal(str(custo)))
                    conn.execute(
                        """UPDATE produtos SET nome=?, marca=?, preco=?, quantidade=?, categoria=?, subcategoria=?,
                               descricao_curta=?, descricao=?, preco_promocional=?, peso=?, largura=?, altura=?,
                               comprimento=?, destaques=?, modo_de_uso=?, lucro=?, ativo=?, rascunho=?, atualizado_em=?
                           WHERE id=?""",
                        (
                            valores["nome"], valores["marca"], valores["preco"], valores["quantidade"], valores["categoria"],
                            valores["subcategoria"], valores["descricao_curta"], valores["descricao"], valores["preco_promocional"],
                            valores["peso"], valores["largura"], valores["altura"], valores["comprimento"], valores["destaques"],
                            valores["modo_de_uso"], lucro, ativo, rascunho, agora_iso, linha.produto_id_existente,
                        ),
                    )
                    produto_id = linha.produto_id_existente
                    registrar_auditoria(conn, "produto", produto_id, "importar_atualizar", usuario=admin_login, depois=valores)
                    atualizados += 1
                else:
                    dados = linha.dados
                    conn.execute(
                        """INSERT INTO produtos (
                            codigo_p, nome, marca, preco, quantidade, categoria, subcategoria, descricao_curta,
                            descricao, preco_promocional, peso, largura, altura, comprimento, destaques, modo_de_uso,
                            custo, lucro, estoque_minimo, ativo, rascunho, atualizado_em
                        ) VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,0,0,0,?,?,?)""",
                        (
                            codigo, dados.get("nome"), dados.get("marca"), dados.get("preco") or 0.0,
                            dados.get("quantidade") or 0, dados.get("categoria"), dados.get("subcategoria"),
                            dados.get("descricao_curta"), dados.get("descricao"), dados.get("preco_promocional"),
                            dados.get("peso"), dados.get("largura"), dados.get("altura"), dados.get("comprimento"),
                            dados.get("destaques"), dados.get("modo_de_uso"), ativo, rascunho, agora_iso,
                        ),
                    )
                    produto_id = int(conn.execute("SELECT last_insert_rowid()").fetchone()[0])
                    registrar_auditoria(conn, "produto", produto_id, "importar_criar", usuario=admin_login, depois=dados)
                    criados += 1

                if linha.tem_imagem and linha.imagem_nome and sessao.zip_resultado:
                    caminho = sessao.zip_resultado.imagens.get(linha.imagem_nome)
                    if caminho:
                        pendentes_imagem.append((produto_id, caminho, caminho.suffix.lower()))
                    else:
                        sem_imagem += 1
                elif linha.classificacao in ("novo", "atualizacao") and not (existente_row and existente_row.get("imagem_url")):
                    sem_imagem += 1
    except sqlite3.IntegrityError as exc:
        falhou = True
        mensagem_erro = "Conflito de dados durante a importação (SKU duplicado detectado no banco)."
        logger.error("importação de produtos falhou por integridade", extra={"evento": "importacao_produtos_integridade"})
    except Exception as exc:  # noqa: BLE001 - erro sanitizado, nunca stack trace ao usuário
        falhou = True
        mensagem_erro = "Falha inesperada ao processar a importação. Nenhum produto foi alterado."
        logger.error("importação de produtos falhou", extra={"evento": "importacao_produtos_falhou", "erro_tipo": type(exc).__name__})

    if falhou:
        _limpar_diretorio(sessao.diretorio_temp)
        _registrar_historico(
            import_id=import_id, admin_login=admin_login, planilha_nome=sessao.planilha_nome, zip_nome=sessao.zip_nome,
            modo=sessao.modo, total=len(sessao.linhas), criados=0, atualizados=0, ignorados=0, com_erro=len(sessao.linhas),
            sem_imagem=0, status="falhou", mensagem_erro=mensagem_erro, inicio=inicio,
        )
        raise ErroImportacao(mensagem_erro or "Falha ao confirmar a importação.")

    imagens_falharam = 0
    for produto_id, caminho, ext in pendentes_imagem:
        try:
            dados_imagem = caminho.read_bytes()
            content_type = {".jpg": "image/jpeg", ".jpeg": "image/jpeg", ".png": "image/png", ".webp": "image/webp"}[ext]
            validar_imagem_real(dados_imagem, content_type)
            normalizada = normalizar_imagem(dados_imagem, content_type)
            resultado = imagem_storage.upload(normalizada, produto_id=str(produto_id), ext=ext, content_type=content_type)
            with conectar() as conn2:
                conn2.execute("UPDATE produtos SET imagem_url=? WHERE id=?", (resultado["url"], produto_id))
        except Exception as exc:  # noqa: BLE001
            imagens_falharam += 1
            sem_imagem += 1
            logger.warning(
                "falha ao promover imagem da importação para o produto %s", produto_id,
                extra={"evento": "importacao_produtos_imagem_falhou", "erro_tipo": type(exc).__name__},
            )

    _limpar_diretorio(sessao.diretorio_temp)
    status_final = "concluido_com_avisos" if (com_erro or ignorados or sem_imagem or imagens_falharam) else "concluido"
    resultado_final = {
        "id": import_id, "total": len(sessao.linhas), "criados": criados, "atualizados": atualizados,
        "ignorados": ignorados, "com_erro": com_erro, "sem_imagem": sem_imagem, "status": status_final,
    }
    _registrar_historico(
        import_id=import_id, admin_login=admin_login, planilha_nome=sessao.planilha_nome, zip_nome=sessao.zip_nome,
        modo=sessao.modo, total=len(sessao.linhas), criados=criados, atualizados=atualizados, ignorados=ignorados,
        com_erro=com_erro, sem_imagem=sem_imagem, status=status_final, mensagem_erro=None, inicio=inicio,
    )
    return resultado_final


def _nome_seguro_arquivo(nome: Optional[str], limite: int = 120) -> Optional[str]:
    if not nome:
        return None
    base = Path(str(nome)).name
    base = re.sub(r"[\r\n\t]", " ", base).strip()
    return base[:limite] or None


def _registrar_historico(*, import_id, admin_login, planilha_nome, zip_nome, modo, total, criados, atualizados,
                          ignorados, com_erro, sem_imagem, status, mensagem_erro, inicio) -> None:
    fim = datetime.now()
    duracao_ms = int((fim - inicio).total_seconds() * 1000)
    try:
        with conectar() as conn:
            conn.execute(
                """INSERT INTO importacoes_produtos (
                    id, iniciado_em, concluido_em, admin_login, planilha_nome, zip_nome, modo, total_linhas,
                    criados, atualizados, ignorados, com_erro, sem_imagem, status, mensagem_erro, duracao_ms
                ) VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)""",
                (
                    import_id, inicio.isoformat(timespec="seconds"), fim.isoformat(timespec="seconds"),
                    _nome_seguro_arquivo(admin_login, 80), _nome_seguro_arquivo(planilha_nome), _nome_seguro_arquivo(zip_nome),
                    modo, total, criados, atualizados, ignorados, com_erro, sem_imagem, status,
                    _texto_seguro(mensagem_erro, limite=200) if mensagem_erro else None, duracao_ms,
                ),
            )
    except Exception:
        logger.error("falha ao registrar histórico de importação", extra={"evento": "importacao_produtos_historico_falhou"})


def listar_historico(limite: int = 20, offset: int = 0) -> list[dict]:
    with conectar() as conn:
        linhas = conn.execute(
            """SELECT id, iniciado_em, concluido_em, admin_login, planilha_nome, zip_nome, modo, total_linhas,
                      criados, atualizados, ignorados, com_erro, sem_imagem, status, mensagem_erro, duracao_ms
               FROM importacoes_produtos ORDER BY iniciado_em DESC LIMIT ? OFFSET ?""",
            (limite, offset),
        ).fetchall()
    return [dict(linha) for linha in linhas]


# --------------------------------------------------------------------------
# Duplicação de produto
# --------------------------------------------------------------------------

CAMPOS_DUPLICAVEIS = [
    "marca", "categoria", "subcategoria", "preco", "custo", "descricao_curta", "descricao",
    "preco_promocional", "peso", "largura", "altura", "comprimento", "destaques", "modo_de_uso",
    "selo", "link_externo", "sob_encomenda", "limite_encomenda",
]


def duplicar_produto(produto_id: int, *, copiar_imagem: bool, admin_login: str) -> dict:
    with conectar() as conn:
        original = conn.execute("SELECT * FROM produtos WHERE id=?", (produto_id,)).fetchone()
        if not original:
            raise ErroImportacao("Produto original não encontrado.")
        original = dict(original)

        novo = {campo: original.get(campo) for campo in CAMPOS_DUPLICAVEIS}
        nome_original = (original.get("nome") or "Produto").strip()
        novo_nome = f"{nome_original} (Cópia)"
        if len(novo_nome) > NOME_MAX:
            novo_nome = novo_nome[: NOME_MAX - len(" (Cópia)")] + " (Cópia)"

        agora_iso = datetime.now().isoformat(timespec="seconds")
        conn.execute(
            """INSERT INTO produtos (
                codigo_p, nome, marca, preco, quantidade, categoria, subcategoria, custo, lucro, estoque_minimo,
                descricao_curta, descricao, preco_promocional, peso, largura, altura, comprimento, destaques,
                modo_de_uso, selo, link_externo, sob_encomenda, limite_encomenda, ativo, rascunho, atualizado_em
            ) VALUES (?,?,?,?,0,?,?,?,?,0,?,?,?,?,?,?,?,?,?,?,?,?,?,0,1,?)""",
            (
                None, novo_nome, novo.get("marca"), novo.get("preco") or 0.0, novo.get("categoria"), novo.get("subcategoria"),
                novo.get("custo") or 0.0, float(Decimal(str(novo.get("preco") or 0.0)) - Decimal(str(novo.get("custo") or 0.0))),
                novo.get("descricao_curta"), novo.get("descricao"), novo.get("preco_promocional"), novo.get("peso"),
                novo.get("largura"), novo.get("altura"), novo.get("comprimento"), novo.get("destaques"), novo.get("modo_de_uso"),
                novo.get("selo"), novo.get("link_externo"), novo.get("sob_encomenda") or 0, novo.get("limite_encomenda") or 10,
                agora_iso,
            ),
        )
        novo_id = int(conn.execute("SELECT last_insert_rowid()").fetchone()[0])

        imagem_copiada = False
        if copiar_imagem and original.get("imagem_url"):
            try:
                if is_managed_by_storage(imagem_storage, original["imagem_url"]):
                    dados = imagem_storage.read(original["imagem_url"])
                    if dados:
                        ext = Path(original["imagem_url"]).suffix.lower() or ".jpg"
                        content_type = {".jpg": "image/jpeg", ".jpeg": "image/jpeg", ".png": "image/png", ".webp": "image/webp"}.get(ext, "image/jpeg")
                        resultado = imagem_storage.upload(dados, produto_id=str(novo_id), ext=ext, content_type=content_type)
                        conn.execute("UPDATE produtos SET imagem_url=? WHERE id=?", (resultado["url"], novo_id))
                        imagem_copiada = True
                else:
                    # URL externa (não gerenciada por este storage, portanto
                    # imutável do nosso ponto de vista): seguro referenciar
                    # diretamente, sem duplicar arquivo.
                    conn.execute("UPDATE produtos SET imagem_url=? WHERE id=?", (original["imagem_url"], novo_id))
                    imagem_copiada = True
            except Exception as exc:  # noqa: BLE001
                logger.warning("falha ao copiar imagem na duplicação de produto", extra={"evento": "duplicar_produto_imagem_falhou", "erro_tipo": type(exc).__name__})

        registrar_auditoria(conn, "produto", novo_id, "duplicar", usuario=admin_login, depois={"origem_id": produto_id, "nome": novo_nome})

    return {"id": novo_id, "nome": novo_nome, "rascunho": True, "imagem_copiada": imagem_copiada, "origem_id": produto_id}
