from __future__ import annotations

import io
from typing import Optional

from fastapi import APIRouter, Cookie, Depends, File, Form, HTTPException, Query, Request, UploadFile
from fastapi.responses import StreamingResponse
from openpyxl import Workbook
from openpyxl.styles import Font
from pydantic import BaseModel, Field

from backend.database import conectar
from backend.logging_config import get_logger
from backend.panel_sessions import exigir_sessao_ou_chave_api
from backend.product_description_templates import gerar_sugestao, listar_modelos
from backend.product_import import (
    CAMPOS_PLANILHA,
    ErroImportacao,
    MARCADOR_LIMPAR,
    MAX_PLANILHA_BYTES,
    MAX_ZIP_BYTES,
    MODOS_VALIDOS,
    confirmar_previa,
    construir_previa,
    criar_previa,
    descartar_previa,
    duplicar_produto,
    ler_planilha,
    limpar_diretorio_temp,
    listar_historico,
    mapear_cabecalhos,
    novo_diretorio_temp,
    obter_previa,
    processar_zip,
)
from backend.rate_limit import limitar_requisicoes

logger = get_logger(__name__)

router = APIRouter(prefix="/api/produtos", tags=["importacao-produtos"])

limitar_modelo = limitar_requisicoes("importar_produtos_modelo", limite=20, janela_segundos=60)
limitar_validar = limitar_requisicoes("importar_produtos_validar", limite=6, janela_segundos=60)
limitar_confirmar = limitar_requisicoes("importar_produtos_confirmar", limite=6, janela_segundos=60)
limitar_historico = limitar_requisicoes("importar_produtos_historico", limite=30, janela_segundos=60)
limitar_duplicar = limitar_requisicoes("duplicar_produto", limite=20, janela_segundos=60)
limitar_modelos_descricao = limitar_requisicoes("modelos_descricao", limite=30, janela_segundos=60)


def _dono_previa(sessao: dict, mistica_painel_sessao: Optional[str]) -> str:
    """Identificador estável da sessão administrativa dona da prévia: o
    token de cookie quando autenticado por sessão de navegador, ou um
    marcador fixo por login quando autenticado por chave de API
    servidor-a-servidor (que não tem um cookie de sessão para vincular)."""
    if mistica_painel_sessao:
        return f"sessao:{mistica_painel_sessao}"
    return f"apikey:{sessao.get('login') or 'integracao'}"


def _admin_dependencia():
    return exigir_sessao_ou_chave_api(perfil_minimo="adm")


def _neutralizar_formula(valor: str) -> str:
    if valor and valor[0] in ("=", "+", "-", "@"):
        return "'" + valor
    return valor


# --------------------------------------------------------------------------
# Planilha-modelo
# --------------------------------------------------------------------------


_CABECALHOS_MODELO = list(CAMPOS_PLANILHA.keys())
_EXEMPLOS_MODELO = [
    {
        "sku": "ESS-LAV-030", "nome": "Essência Lavanda 30ml", "categoria": "Aromaterapia",
        "subcategoria": "Essências", "marca": "Via Aroma", "descricao_curta": "Essência de lavanda para aromatizador",
        "descricao": "Essência concentrada de lavanda, ideal para aromatizadores elétricos.",
        "preco": "34,90", "preco_promocional": "", "estoque": "20", "peso": "0.05", "largura": "4",
        "altura": "9", "comprimento": "4", "destaques": "Fragrância relaxante", "modo_de_uso": "",
        "imagem": "essencia-lavanda.jpg", "ativo": "não", "status": "rascunho",
    },
    {
        "sku": "DIF-CIT-120", "nome": "Difusor de Varetas Citronela 120ml", "categoria": "Aromaterapia",
        "subcategoria": "Difusores", "marca": "Via Aroma", "descricao_curta": "Difusor de varetas citronela",
        "descricao": "Difusor de varetas com fragrância de citronela, perfuma o ambiente por semanas.",
        "preco": "49,90", "preco_promocional": "39,90", "estoque": "15", "peso": "0.2", "largura": "6",
        "altura": "15", "comprimento": "6", "destaques": "Repele insetos;Longa duração", "modo_de_uso": "",
        "imagem": "difusor-citronela.jpg", "ativo": "não", "status": "rascunho",
    },
]

_ORIENTACOES = [
    ("Como funciona a importação", ""),
    ("1. Preencha esta planilha", "Use exatamente os nomes de coluna da primeira linha da aba 'produtos'. Não renomeie nem reordene é permitido, mas colunas desconhecidas são ignoradas com aviso."),
    ("2. Valide antes de confirmar", "No painel, clique em 'Validar e visualizar' para ver uma prévia. Nada é gravado nesta etapa."),
    ("3. Confirme explicitamente", "Só depois de revisar a prévia, clique em 'Confirmar importação' para gravar os produtos."),
    ("", ""),
    ("Campos obrigatórios", "nome (sempre). sku é obrigatório apenas para ATUALIZAR um produto existente."),
    ("Campos opcionais", "categoria, subcategoria, marca, descricao_curta, descricao, preco_promocional, peso, largura, altura, comprimento, destaques, modo_de_uso, imagem, ativo, status."),
    ("", ""),
    ("Formato de preco/preco_promocional", "Número com vírgula ou ponto decimal, ex.: 34,90 ou 34.90. Deixe vazio para não definir."),
    ("Formato de estoque", "Número inteiro (0 a 1.000.000)."),
    ("Formato de peso/largura/altura/comprimento", "Número decimal (kg e cm), opcional."),
    ("Valores aceitos para 'ativo'", "sim / não / verdadeiro / falso / 1 / 0."),
    ("Valores aceitos para 'status'", "ativo / inativo / rascunho (tem prioridade sobre a coluna 'ativo' quando preenchido)."),
    ("", ""),
    ("Modos de importação", ""),
    ("Criar somente novos como rascunho (padrão)", "Produtos com SKU já existente são ignorados; nenhum produto é ativado automaticamente."),
    ("Criar somente novos como ativos", "Cria só produtos novos, já publicados; exige confirmação extra no painel."),
    ("Criar novos e atualizar existentes pelo SKU", "Requer SKU para atualizar; célula vazia preserva o valor atual do produto."),
    ("Criar e atualizar pelo SKU, todos como rascunho", "Novos e atualizados sempre viram rascunho após a importação."),
    ("", ""),
    ("Regra de correspondência de imagem", "A coluna 'imagem' deve conter o nome exato do arquivo dentro do ZIP (ex.: produto-1.jpg). Maiúsculas/minúsculas e espaços extras são ignorados na comparação."),
    ("Célula vazia em produto existente", f"Nunca apaga o valor atual. Para apagar um campo opcional de propósito, use o marcador {MARCADOR_LIMPAR} (não funciona em nome, categoria, sku nem imagem)."),
    ("Imagem vazia", "Nunca remove a imagem atual de um produto existente."),
    ("Aviso sobre fórmulas", "Este arquivo não deve conter fórmulas. Valores iniciados por =, +, - ou @ são tratados como texto, nunca executados."),
    ("Aviso sobre SKU na atualização", "Um SKU já usado por outro produto (inclusive inativo) nunca é reatribuído; a linha é rejeitada com erro."),
]


def _gerar_planilha_modelo() -> bytes:
    wb = Workbook()
    aba = wb.active
    aba.title = "produtos"
    negrito = Font(bold=True)
    for coluna, cabecalho in enumerate(_CABECALHOS_MODELO, start=1):
        celula = aba.cell(row=1, column=coluna, value=cabecalho)
        celula.font = negrito
    for linha_idx, exemplo in enumerate(_EXEMPLOS_MODELO, start=2):
        for coluna, cabecalho in enumerate(_CABECALHOS_MODELO, start=1):
            valor = str(exemplo.get(cabecalho, ""))
            aba.cell(row=linha_idx, column=coluna, value=_neutralizar_formula(valor))

    orientacoes = wb.create_sheet("orientacoes")
    orientacoes.cell(row=1, column=1, value="Orientações de preenchimento").font = Font(bold=True, size=13)
    linha_idx = 3
    for titulo, texto in _ORIENTACOES:
        if titulo:
            celula = orientacoes.cell(row=linha_idx, column=1, value=_neutralizar_formula(titulo))
            if not texto:
                celula.font = Font(bold=True)
        if texto:
            orientacoes.cell(row=linha_idx, column=2, value=_neutralizar_formula(texto))
        linha_idx += 1
    orientacoes.column_dimensions["A"].width = 42
    orientacoes.column_dimensions["B"].width = 90

    buffer = io.BytesIO()
    wb.save(buffer)
    return buffer.getvalue()


@router.get("/importacao/modelo", dependencies=[Depends(limitar_modelo)])
def baixar_planilha_modelo(sessao: dict = Depends(_admin_dependencia())):
    conteudo = _gerar_planilha_modelo()
    return StreamingResponse(
        io.BytesIO(conteudo),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=modelo-importacao-produtos.xlsx"},
    )


# --------------------------------------------------------------------------
# Validação / prévia
# --------------------------------------------------------------------------


async def _ler_upload_limitado(arquivo: UploadFile, limite_bytes: int, *, nome_campo: str) -> bytes:
    dados = bytearray()
    while True:
        pedaco = await arquivo.read(1024 * 256)
        if not pedaco:
            break
        dados.extend(pedaco)
        if len(dados) > limite_bytes:
            raise HTTPException(status_code=413, detail=f"{nome_campo} excede o tamanho máximo permitido.")
    return bytes(dados)


@router.post("/importacao/validar", dependencies=[Depends(limitar_validar)])
async def validar_planilha_importacao(
    request: Request,
    planilha: UploadFile = File(...),
    zip_imagens: Optional[UploadFile] = File(None),
    modo: str = Form(...),
    sessao: dict = Depends(_admin_dependencia()),
    mistica_painel_sessao: Optional[str] = Cookie(default=None),
):
    if modo not in MODOS_VALIDOS:
        raise HTTPException(status_code=400, detail="Modo de importação inválido.")

    conteudo_planilha = await _ler_upload_limitado(planilha, MAX_PLANILHA_BYTES, nome_campo="A planilha")
    try:
        linhas_brutas = ler_planilha(planilha.filename or "planilha", conteudo_planilha)
    except ErroImportacao as exc:
        raise HTTPException(status_code=400, detail=str(exc))

    try:
        mapa_cabecalhos, colunas_desconhecidas = mapear_cabecalhos(linhas_brutas[0])
    except ErroImportacao as exc:
        raise HTTPException(status_code=400, detail=str(exc))

    diretorio_temp = novo_diretorio_temp()
    zip_resultado = None
    zip_nome = None
    try:
        if zip_imagens is not None and zip_imagens.filename:
            conteudo_zip = await _ler_upload_limitado(zip_imagens, MAX_ZIP_BYTES, nome_campo="O arquivo ZIP")
            try:
                zip_resultado = processar_zip(conteudo_zip, diretorio_temp)
            except ErroImportacao as exc:
                raise HTTPException(status_code=400, detail=str(exc))
            zip_nome = zip_imagens.filename

        with conectar() as conn:
            linhas, resumo = construir_previa(conn, linhas_brutas, mapa_cabecalhos, modo, zip_resultado)
    except HTTPException:
        limpar_diretorio_temp(diretorio_temp)
        raise

    resumo["colunas_desconhecidas"] = colunas_desconhecidas

    dono = _dono_previa(sessao, mistica_painel_sessao)
    previa = criar_previa(
        dono=dono, admin_login=sessao.get("login") or "adm", modo=modo,
        planilha_nome=planilha.filename or "planilha", zip_nome=zip_nome,
        linhas=linhas, resumo=resumo, diretorio_temp=diretorio_temp, zip_resultado=zip_resultado,
    )

    return {
        "token": previa.token,
        "expira_em_segundos": int(previa.expira_em - previa.criado_em),
        "modo": modo,
        "resumo": resumo,
        "linhas": _serializar_linhas(linhas),
    }


def _serializar_linhas(linhas) -> list[dict]:
    return [
        {
            "numero": l.numero,
            "classificacao": l.classificacao,
            "sku": l.dados.get("codigo_p"),
            "nome": l.dados.get("nome"),
            "tem_imagem": l.tem_imagem,
            "erros": l.erros,
            "avisos": l.avisos,
        }
        for l in linhas
    ]


@router.get("/importacao/previews/{token}", dependencies=[Depends(limitar_historico)])
def obter_preview_importacao(
    token: str,
    status: Optional[str] = Query(default=None, description="novo|atualizacao|ignorado|erro"),
    pagina: int = Query(default=1, ge=1),
    tamanho_pagina: int = Query(default=50, ge=1, le=200),
    sessao: dict = Depends(_admin_dependencia()),
    mistica_painel_sessao: Optional[str] = Cookie(default=None),
):
    dono = _dono_previa(sessao, mistica_painel_sessao)
    try:
        previa = obter_previa(token, dono=dono)
    except ErroImportacao as exc:
        raise HTTPException(status_code=404, detail=str(exc))

    linhas = previa.linhas
    if status:
        linhas = [l for l in linhas if l.classificacao == status]
    inicio = (pagina - 1) * tamanho_pagina
    pagina_linhas = linhas[inicio: inicio + tamanho_pagina]

    return {
        "token": token,
        "resumo": previa.resumo,
        "total_filtrado": len(linhas),
        "pagina": pagina,
        "tamanho_pagina": tamanho_pagina,
        "linhas": _serializar_linhas(pagina_linhas),
        "confirmado": previa.confirmado,
    }


class ConfirmarImportacaoIn(BaseModel):
    token: str = Field(min_length=1, max_length=200)


@router.post("/importacao/confirmar", dependencies=[Depends(limitar_confirmar)])
def confirmar_importacao(
    payload: ConfirmarImportacaoIn,
    sessao: dict = Depends(_admin_dependencia()),
    mistica_painel_sessao: Optional[str] = Cookie(default=None),
):
    dono = _dono_previa(sessao, mistica_painel_sessao)
    try:
        resultado = confirmar_previa(payload.token, dono=dono, admin_login=sessao.get("login") or "adm")
    except ErroImportacao as exc:
        raise HTTPException(status_code=400, detail=str(exc))
    return resultado


@router.delete("/importacao/previews/{token}", dependencies=[Depends(limitar_historico)])
def cancelar_previa_importacao(
    token: str,
    sessao: dict = Depends(_admin_dependencia()),
    mistica_painel_sessao: Optional[str] = Cookie(default=None),
):
    dono = _dono_previa(sessao, mistica_painel_sessao)
    try:
        obter_previa(token, dono=dono)
    except ErroImportacao as exc:
        raise HTTPException(status_code=404, detail=str(exc))
    descartar_previa(token)
    return {"ok": True}


@router.get("/importacao/historico", dependencies=[Depends(limitar_historico)])
def historico_importacoes(
    pagina: int = Query(default=1, ge=1),
    tamanho_pagina: int = Query(default=20, ge=1, le=100),
    sessao: dict = Depends(_admin_dependencia()),
):
    offset = (pagina - 1) * tamanho_pagina
    itens = listar_historico(limite=tamanho_pagina, offset=offset)
    return {"pagina": pagina, "tamanho_pagina": tamanho_pagina, "itens": itens}


# --------------------------------------------------------------------------
# Duplicação de produto
# --------------------------------------------------------------------------


class DuplicarProdutoIn(BaseModel):
    copiar_imagem: bool = False


@router.post("/{produto_id}/duplicar", dependencies=[Depends(limitar_duplicar)])
def duplicar_produto_endpoint(
    produto_id: int,
    payload: DuplicarProdutoIn,
    sessao: dict = Depends(_admin_dependencia()),
):
    try:
        resultado = duplicar_produto(produto_id, copiar_imagem=payload.copiar_imagem, admin_login=sessao.get("login") or "adm")
    except ErroImportacao as exc:
        raise HTTPException(status_code=404, detail=str(exc))
    return resultado


# --------------------------------------------------------------------------
# Modelos de descrição
# --------------------------------------------------------------------------


@router.get("/modelos-descricao", dependencies=[Depends(limitar_modelos_descricao)])
def listar_modelos_descricao_endpoint(sessao: dict = Depends(_admin_dependencia())):
    return {"modelos": listar_modelos()}


class GerarModeloDescricaoIn(BaseModel):
    categoria: str = Field(min_length=1, max_length=40)
    campos: dict = Field(default_factory=dict)


@router.post("/modelos-descricao/gerar", dependencies=[Depends(limitar_modelos_descricao)])
def gerar_modelo_descricao_endpoint(payload: GerarModeloDescricaoIn, sessao: dict = Depends(_admin_dependencia())):
    try:
        sugestao = gerar_sugestao(payload.categoria, payload.campos)
    except KeyError:
        raise HTTPException(status_code=404, detail="Modelo de descrição não encontrado.")
    return sugestao
